"""Read salary.xlsx and dump structure + first rows."""
import openpyxl, csv, sys

path = r"c:\xampp\htdocs\projectx\watanybot\sources\primary\salaries\salary.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)

for name in wb.sheetnames:
    ws = wb[name]
    print(f"=== Sheet: {name} | Rows: {ws.max_row} | Cols: {ws.max_column} ===")
    for row_idx in range(1, min(8, ws.max_row + 1)):
        cells = []
        for col_idx in range(1, ws.max_column + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            cells.append(str(v) if v is not None else "")
        print(f"  R{row_idx}: " + " | ".join(cells))
    print()

# Export first sheet as full CSV
ws = wb[wb.sheetnames[0]]
out = r"c:\xampp\htdocs\projectx\watanybot\sources\primary\salaries\salary_full.csv"
with open(out, "w", encoding="utf-8", newline="") as f:
    writer = csv.writer(f)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=True):
        writer.writerow([str(v) if v is not None else "" for v in row])
print(f"Exported {ws.max_row} rows to {out}")
